import pyautogui
from openpyxl import load_workbook
import time

workbook = load_workbook(
    "C:\\Users\\t0268052\\Desktop\\Mass Port-Down\\Duplicate - 2\\TCIS Duplicates Cleanup - Todd.xlsx")
worksheet = workbook.active

Group_list = []  # Here we store all the Group_list records
extract_list = []  # Here we store all the 54xxxx and other part IDS

i = 189  # set to begin at row 2

while worksheet.cell(row=i, column=1).value is not None:  # Set column 1
    if i != 196 and i != 218 and i != 219:  # These are the exceptions
        Group_list.append(worksheet.cell(row=i, column=1).value)
        i += 1
    else:
        i += 1

i = 189  # reset i to row 2

while worksheet.cell(row=i, column=2).value is not None:  # Set column 2
    if i != 196 and i != 218 and i != 219:  # These are the exceptions
        extract_list.append(worksheet.cell(row=i, column=2).value)
        i += 1
    else:
        i += 1


def desc_finder(index):                             # Given the index of the group list, it will determine whether
    print(extract_list[index][0])                   # the part ID is recent or old. If Ture its new, otherwise its old.
    if extract_list[index][0] == "5":
        return True

    else:
        return False

time.sleep(3)
pyautogui.click(146, 367)                           # Click on MM02 twice to open
pyautogui.click(146, 367)
time.sleep(0.3)

x = 0
length = len(Group_list)

while length is not 0:                      # Until the parts are finished, heep updating comments
    time.sleep(0.5)
    pyautogui.typewrite(Group_list[x])
    time.sleep(0.5)
    pyautogui.typewrite(["enter"])
    time.sleep(0.5)
    pyautogui.press("pagedown")
    pyautogui.click(528, 762)
    time.sleep(0.5)
    pyautogui.click(605, 255)
    time.sleep(0.5)
    pyautogui.click(365, 531)
    time.sleep(0.5)
    value = desc_finder(x)

    if value:                                                                #  If true, put the new comment
        pyautogui.typewrite("November 28, 2022 - Grigor Pahlevanyan:")
        pyautogui.typewrite(["enter"])
        pyautogui.typewrite("Part re-ported from TCIS under new Group BCN Part ID")
        pyautogui.typewrite(["enter"])
        pyautogui.typewrite("due to duplicate part issue from Lot 3.")
        pyautogui.typewrite(["enter"])
        pyautogui.typewrite(["enter"])

        #input("if you enter something it will continue: ")
        time.sleep(3)
        pyautogui.click(376,80)                                             # press save

    else:                                                                   #  If false, put the old comment
        pyautogui.typewrite("November 28, 2022 - Grigor Pahlevanyan:")
        pyautogui.typewrite(["enter"])
        pyautogui.typewrite("Uploaded to TCIS under Lot 3. Linked back down to SAP")
        pyautogui.typewrite(["enter"])
        pyautogui.typewrite("and verified. CECN not required.")
        pyautogui.typewrite(["enter"])
        pyautogui.typewrite(["enter"])

        #input("if you enter something it will continue: ")
        time.sleep(3)
        pyautogui.click(376,80)                                             # press save

    time.sleep(0.5)
    pyautogui.click(146, 367)                                               # Click the back arrow
    time.sleep(0.5)
    x += 1
    length -= 1


#scroll down in MM02 Point(x=1158, y=949)

# print(Group_list)
# print(extract_list)

print(pyautogui.position())
