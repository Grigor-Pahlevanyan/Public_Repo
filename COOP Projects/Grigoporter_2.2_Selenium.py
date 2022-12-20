import sys
from openpyxl import load_workbook  # For excel
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time

PATH = r"C:\Users\t0268052\PycharmProjects\Porting_Automation\chromedriver.exe"  # the r means "raw string"
driver = webdriver.Chrome(PATH)                  # This is how the program utilizes the web driver

# Below is the link to the website we want to open.
driver.get("https://tcis.online.corp.thales/wps/myportal/!ut/p/a1/hc7NCsIwEATgR8qkP2lzXFuaRA1CIFhzkZxKQasH8fmN0psY9zbwDbMssJGFJT7nKT7m2xIv7xzEWVJnlK74VtUDgdodPKwrYXkCpwTw4wi5vpLF2s-AP_tHFj6kU6SrZp82q7aA6Te6b6QFjFhB7sX8yBdA7TlIkTo450tAsPvVj5jN9AJeQL9t/dl5/d5/L2dJQSEvUUt3QS80SmlFL1o2XzlBQ0lHSDQxSkc1RkEwQThLMFUwTVIzME0x/#")

login = input("Please log in to TCIS. Enter Y/N once logged in: ")

if login == "N" or login == "n":
    driver.close()                              # close() will close the tab, quit() will close the browser.
    sys.exit()                                  # no log in, leave the program

# For batches
#  workbook = load_workbook('C:\\Users\\t0268052\\Desktop\\Mass Port-Down\\Batch 13\\Grigor - Batch.xlsx')

# For CECNs
workbook = load_workbook("C:\\Users\\t0268052\\Desktop\\CECN\\CCP00499340 - Start Nov 9\\CCP00499340 Descriptions.xlsx")
worksheet = workbook.active

Group_list = []                          # Here we store all the Group_list records
Desc_list = []                           # Here we store all the descriptions corresponding to the Group_list # records
i = 11                                    # set to begin at row 2

while worksheet.cell(row=i, column=2).value is not None:        # Set column 1
    Group_list.append(worksheet.cell(row=i, column=2).value)
    i = i + 1

i = 11                                    # reset i to row 2

while worksheet.cell(row=i, column=8).value is not None:        # Set column 5
    Desc_list.append(worksheet.cell(row=i, column=8).value)
    i = i + 1

Group_size = len(Group_list)              # Updated from Group_size = i - 2 (because you might start at midway of Excel)
print("# of ROH: ", Group_size)

# printing the ROHs
k = 0
while k < Group_size:
    print(Group_list[k])
    k += 1

Group_count = 0                          # change back to 0********
while Group_count < Group_size:
    search = driver.find_element(By.NAME, "partId")
    search.clear()
    time.sleep(1)
    search.send_keys(Group_list[Group_count])  # Place the Group_list numbers from the Batches
    search.send_keys(Keys.RETURN)

    # --------------- This section is to identify how many times to loop -----------------------------
    try:
        element = WebDriverWait(driver, 20).until(
            ec.presence_of_element_located((By.ID, "t3_link"))  # Waiting for the website to load
        )
    finally:
        time.sleep(3)
        search = driver.find_element(By.ID, "t3_link").click()  # Click Sources TAB
        time.sleep(3)

    Source_count = len(driver.find_elements(By.ID, "col4")) - 1  # make a list of all elements
    # _____________________________________________________________________________________^

    # Determining whether a part "Group/source" is selected or not______________________________________
    images = driver.find_elements(By.TAG_NAME, "img")
    selection = []  # Which sources are selected and which are not, in respective order
    counter = 0
    print("\n")
    for image in images:
        source_img = image.get_attribute("title")
        if source_img == "Selected by the Operational Unit":
            print(source_img)
            selection.append(True)                                      # if selected, then append a true
            counter += 1
        elif source_img == "Not Selected by the Operational Unit":
            print(source_img)
            selection.append(False)                                     # if not selected, then append a false
            counter += 1
    print("\n")
    print("Total number of selections including the Group", counter)
    print("source count", Source_count)
    # ___________________________________________________________________________________________________^

    # the code below insures the group that is selected has a green check mark__________________________________
    if counter - 1 != Source_count:  # This ensures the counter is strictly greater than the source count by 1
        print("the program will pause until the group is linked: ")
        user_imp = input("Please enter 'Fixed' if the group is selected: ")
        if user_imp != "Fixed" or user_imp != "fixed":       # if the group isn't linked then quit the program
            driver.close()                           # close() will close the tab, quit() will close the browser.
            sys.exit()                               # no log in, leave the program

    print(selection)                                 # Prints our list of selected/unselected parts
    print("\n")
    # __________________________________________________________________________________________________________^

    # The code below updates the selected part ______________________________________________
    def updater(description_list):
        update = driver.find_elements(By.CLASS_NAME, "button")
        update[3].click()                                            # clicks on the update button
        time.sleep(1)
        unit_data = driver.find_elements(By.TAG_NAME, "a")
        unit_data[1].click()                                         # clicks on "update unit data"

        desc_extract = driver.find_element(By.TAG_NAME, "textarea").get_attribute("textContent")
        print("\t", desc_extract)

        if "*" not in desc_extract:
            print("\tNot a German description\n")
            desc = driver.find_element(By.NAME, "UNIT_ATTRIBUTE_DESCRIPTION")
            desc.clear()  # Clear the text inside the box
            time.sleep(1)
            desc.send_keys(Desc_list[Group_count])  # Place the description corresponding to that group
            time.sleep(1)

        else:
            print("\tGerman description, will leave it as is\n")

        TSD_bar = driver.find_elements(By.TAG_NAME, "select")
        TSD_bar[4].click()                                          # Clicks the "SAP Authorization Group" tab
        time.sleep(1)
        TSD_bar[4].send_keys("T")                                   # Enters T
        time.sleep(0.5)
        TSD_bar[4].send_keys(Keys.RETURN)                           # Once entered T, hits enter key for "TSD"
        time.sleep(0.5)
        update_part = driver.find_elements(By.CLASS_NAME, "button")
        time.sleep(1)
        update_part[1].click()                                      # Clicks to update the part
        time.sleep(1)
        search_tab = driver.find_elements(By.TAG_NAME, "a")
        time.sleep(0.5)
        search_tab[49].click()                                      # Go back to the search
        return None

    updater(Desc_list[Group_count])                                 # Calling the function

    '''
    a = 0
    for button in search_tab:
        description = button.get_attribute("href")
        print(a, "Description: ", description)
        a += 1
    '''
    # ______________________________________________________________________________________^

    # Updating the sources__________________________________________________________________
    # Assuming Source_list[0] is already updated
    time.sleep(2)
    count = 1                                                       # Starting from the first source at index 1
    while count <= Source_count:
        print(count, "Source Loop")
        try:
            element = WebDriverWait(driver, 20).until(
                ec.presence_of_element_located((By.ID, "t3_link"))  # Waiting for the website to load
            )
        finally:
            # Search the group number (ROH in SAP terms)
            print("\tResearch group number")
            search = driver.find_element(By.NAME, "partId")
            search.clear()
            time.sleep(1)
            search.send_keys(Group_list[Group_count])                # Place the Group_list numbers from the Batches
            search.send_keys(Keys.RETURN)

        try:
            element = WebDriverWait(driver, 20).until(
                ec.presence_of_element_located((By.ID, "t3_link"))  # Waiting for the website to load
            )
        finally:
            # Click the sources TAB
            time.sleep(10)
            print("\tClicking sources TAB")
            search = driver.find_element(By.ID, "t3_link").click()  # click sources again

        try:
            element = WebDriverWait(driver, 20).until(
                ec.presence_of_element_located((By.ID, "t3_link"))  # Waiting for the website to load
            )
        finally:
           # Clicking on the individual source (HERS in SAP terms)
            Source_list = driver.find_elements(By.ID, "col4")
            time.sleep(10)
            if selection[count]:                                    # Click the source, only if it is valid
                print("\tClicking a source")
                Source_list[count].click()                          # Click one of the sources
                time.sleep(10)
                print("\tUpdate the source here")

                # The code below updates the selected part ______________________________________________
                updater(Desc_list[Group_count])                     # Calling the function that updates the selected
                # ______________________________________________________________________________________^

            count += 1                                              # increment the source counter
    Group_count += 1                                                # increment the group counter

# The code below will close the program is needed by the user
leave = input("Finished, do you want to quit? (Enter Y/N): ")

if leave == "Y" or leave == "y":
    driver.close()                                       # close() will close the tab, quit() will close the browser.
    sys.exit()                                           # no log in, leave the program


"""
Future Updates: Make it work with Toronto local numbers and custom 90XXXX numbers.
In the update part section they have an extra input box, so you need to adjust where you clock for the "TSD" bar.
So make a condition, if the ROH is not Thales Global 99* then use the other location for clicking.
"""