import pyautogui
from openpyxl import load_workbook
import time

# Batches Below
workbook = load_workbook("C:\\Users\\t0268052\\Desktop\\Mass Port-Down\\Batch 14-16\\Batch 14-16.xlsx")

# CECNs Below
#workbook = load_workbook("C:\\Users\\t0268052\\Desktop\\CECN\\CCP00500409 - Start Nov 11\\CCP00500409 Descriptions.xlsx")

worksheet = workbook.active

Group_list = []  # Here we store all the Group_list records

i = 23  # set to begin at row 2

while worksheet.cell(row=i, column=1).value is not None:  # Set column 1
    Group_list.append(worksheet.cell(row=i, column=1).value)
    i += 1

print(Group_list)

time.sleep(3)
pyautogui.click(146, 367)                           # Click on MM02 twice to open
pyautogui.click(146, 367)
time.sleep(0.3)

x = 0
length = len(Group_list)
print(length)

while x is not length:                      # Until the parts are finished, keep updating comments
    time.sleep(0.5)
    a = Group_list[x]
    a = str(a)                              # For some reason, pyautogui.typewrite() only takes strings.
    time.sleep(0.5)
    pyautogui.typewrite(a)
    time.sleep(0.5)
    pyautogui.typewrite(["enter"])          # Put the Group Number in
    time.sleep(0.5)
    pyautogui.click(255, 251)               # Click on "Basic Data 2"
    time.sleep(0.5)
    pyautogui.click(362, 430)               # Click on "Prod./Insp. memo"
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')           # Override the old content with the below
    pyautogui.typewrite('    #2     ROHS')
    time.sleep(0.5)
    pyautogui.click(1164, 253)              # Click the box to open all sections
    time.sleep(0.5)
    pyautogui.click(1234, 431)              # Choose "Purchasing" from the field
    time.sleep(0.5)
    pyautogui.click(345, 492)               # Click on "Purchasing Group"
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')           # Override the old content with "517"
    pyautogui.typewrite('517')
    time.sleep(0.5)
    pyautogui.click(1164, 253)              # again, click the box to open all the sections
    time.sleep(0.5)
    pyautogui.click(1281, 798)              # Click on "Quality Management"
    time.sleep(0.5)
    pyautogui.click(349, 760)               # Click on "Certificate Type"
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')           # Override the old content with "NONE"
    pyautogui.typewrite('NONE')
    time.sleep(0.5)
    pyautogui.typewrite(["enter"])          # Click enter twice to move to the next page (Accounting 1)
    time.sleep(0.5)
    pyautogui.typewrite(["enter"])
    time.sleep(0.5)
    pyautogui.click(914, 700)               # Click on the "Standard Price"
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')           # Override the old content with "0.01"
    pyautogui.typewrite('0.01')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 's')           # SAVE
    time.sleep(0.5)

    x += 1
    #length -= 1